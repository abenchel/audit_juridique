"""Convertit EnerVivo_Liste_Documents_Projet_V2.xlsx → documents_projet_v2.json.

V2 = référentiel d'**enrichissement** par dossier (N1/N2/N3), pas par jalon.
Pour chaque type de document il dit :
  - où il se trouve dans l'arborescence projet (dossier_n1 + sous_dossier)
  - quelles extensions sont attendues
  - un commentaire (souvent : conditionnel ou indice de jalon)

Sortie : config/documents_projet_v2.json structuré :
{
  "version": "V2",
  "source": "EnerVivo_Liste_Documents_Projet_V2.xlsx",
  "total_documents": 227,
  "documents": [
    {
      "code": "promesse_de_bail_signee",
      "name": "Promesse de Bail signee",
      "folder_n1": "4-Documents Administratifs",
      "folder_path": "4-Documents Administratifs/6-Bail/PDB",
      "extensions": [".pdf"],
      "comment": "PDB notariee ou sous seing prive"
    }
  ]
}

Usage : python -m scripts.convert_liste_projet_to_json \
            --in EnerVivo_Liste_Documents_Projet_V2.xlsx \
            --out config/documents_projet_v2.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl

SHEET_NAME = "Liste exhaustive documents"
HEADER_ROW = 3  # 0-indexed, ligne d'en-tête "Dossier (N1) | Sous-dossier ..."


def slugify(value: str) -> str:
    nfkd = unicodedata.normalize("NFKD", value)
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_str).strip("_").lower()
    return slug or "doc"


def _parse_extensions(raw: str | None) -> list[str]:
    """Parse '.pdf, .docx' ou '.pdf' ou 'tous formats' → liste normalisée."""
    if not raw:
        return []
    s = str(raw).lower().strip()
    if "tous" in s:
        return []  # joker : toutes extensions acceptées
    # Split sur virgule, slash, espace multiple
    parts = re.split(r"[,/\s]+", s)
    exts: list[str] = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if not p.startswith("."):
            p = "." + p
        if p not in exts:
            exts.append(p)
    return exts


def _build_folder_path(n1: str, n2: str | None) -> str:
    """Concatène N1 + N2 en remplaçant '>' par '/'. '(racine)' → ignoré."""
    n1 = (n1 or "").strip()
    n2 = (n2 or "").strip()
    if not n2 or n2 == "(racine)":
        return n1
    # N2 utilise '>' comme séparateur de sous-niveau, on normalise en '/'
    n2_clean = re.sub(r"\s*>\s*", "/", n2).strip().strip("/")
    return f"{n1}/{n2_clean}"


def convert(input_path: Path, output_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Feuille '{SHEET_NAME}' absente. Sheets : {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    seen_slugs: dict[str, int] = {}
    docs: list[dict[str, Any]] = []

    # Colonnes : A=N1, B=N2, C=Type, D=Extensions, E=Commentaire
    for row in ws.iter_rows(min_row=HEADER_ROW + 2, values_only=True):
        if not row or len(row) < 3:
            continue
        n1, n2, type_doc, ext_raw, comment = (
            row[0], row[1], row[2], row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
        )

        # Skip lignes section (N1 seul, type vide) et lignes vides
        if not type_doc:
            continue
        type_doc = str(type_doc).strip()
        # Header parasitaire qui peut leaker si HEADER_ROW est mal aligné
        if type_doc.lower() in ("type de document", "type"):
            continue

        n1 = str(n1).strip() if n1 else ""
        if not n1:
            # Ligne sans N1 → on ignore (probablement un sous-titre orphelin)
            continue

        folder_path = _build_folder_path(n1, n2)
        extensions = _parse_extensions(ext_raw)

        base = slugify(type_doc)
        count = seen_slugs.get(base, 0) + 1
        seen_slugs[base] = count
        code = base if count == 1 else f"{base}_{count}"

        docs.append(
            {
                "code": code,
                "name": type_doc,
                "folder_n1": n1,
                "folder_path": folder_path,
                "extensions": extensions,
                "comment": (str(comment).strip() if comment else "") or None,
            }
        )

    result = {
        "version": "V2",
        "source": input_path.name,
        "total_documents": len(docs),
        "documents": docs,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Convertit V2.xlsx → documents_projet_v2.json")
    parser.add_argument("--in", dest="input", required=True, help="Chemin V2.xlsx")
    parser.add_argument("--out", dest="output", default="config/documents_projet_v2.json")
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    if not in_path.exists():
        print(f"Introuvable : {in_path}", file=sys.stderr)
        sys.exit(1)

    result = convert(in_path, out_path)
    print(f"{result['total_documents']} documents convertis")
    print(f"Ecrit : {out_path}")
    # Répartition par N1
    by_n1: dict[str, int] = {}
    for d in result["documents"]:
        by_n1[d["folder_n1"]] = by_n1.get(d["folder_n1"], 0) + 1
    for n1, c in sorted(by_n1.items()):
        print(f"  {n1:>35} : {c} docs")


if __name__ == "__main__":
    main()
