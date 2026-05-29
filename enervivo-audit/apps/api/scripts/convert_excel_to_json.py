"""Convertit le référentiel V12 (Excel EnerVivo) en documents_v12.json.

Source : 260518_Document_par_Jalon_V12.xlsx, feuille 'Liste documents par jalon'.

Colonnes V12 attendues (ligne 4) :
  # | Jalon | Document | PROPRIETE | Type de versioning | Jalons concernes |
  Description | Format de doc | Lien_DIBOS_H | Lien_DMONFLANQUIN

Sortie : config/documents_v12.json structuré par jalon, avec :
  - code      : slug stable (ex. 'j1_pdb_signee')
  - name      : libellé original
  - jalon     : 'Avant J1', 'J1', ..., 'J5_Construction', 'J6_MES', 'J7_Cloture'
  - propriete : 'Obligatoire' | 'Facultatif' | 'Cas par cas' | 'Annexes 3 PDB'
  - versioning : libellé du type de versioning
  - jalons_concernes : libellé tel quel (souvent égal à `jalon`)
  - note      : description V12 (1 phrase métier — quand le doc s'applique, sa
                définition, ses pièges). Champ nommé `note` pour compat aval.
  - format    : extension(s) attendue(s) (ex. '.pdf', '.xlsx', '.pdf, .docx')

Vérité-terrain V12 (colonnes Lien_DIBOS_H / Lien_DMONFLANQUIN) : NON exportée
ici — elle sert hors pipeline pour évaluer la qualité de l'audit.

Gestion des doublons de # : certains documents partagent le même # (alternatives,
ex. #17 personne physique vs morale). Le slug est dédupliqué par suffixe _2, _3.

Usage : python -m scripts.convert_excel_to_json --in <path.xlsx> --out config/documents_v12.json
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl

SHEET_NAME = "Liste documents par jalon"
HEADER_ROW = 3  # 0-indexed → données à partir de la ligne 5 (1-indexed)
JALONS_ORDER = [
    "Avant J1",
    "J1",
    "J2a",
    "J2b",
    "J3",
    "J4",
    "J5_Construction",
    "J6_MES",
    "J7_Cloture",
]


def slugify(value: str) -> str:
    """Slug ASCII stable (sans dépendance externe)."""
    nfkd = unicodedata.normalize("NFKD", value)
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_str).strip("_").lower()
    return slug or "doc"


def convert(input_path: Path, output_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"❌ Feuille '{SHEET_NAME}' absente. Sheets : {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    seen_slugs: dict[str, int] = {}
    docs_by_jalon: dict[str, list[dict[str, Any]]] = {}

    for row in ws.iter_rows(min_row=HEADER_ROW + 2, values_only=True):
        # V12 : 10 colonnes — # | Jalon | Document | PROPRIETE | Versioning |
        # Jalons concernes | Description | Format | Lien_DIBOS_H | Lien_DMONFLANQUIN
        num = row[0]
        jalon = row[1]
        document = row[2]
        propriete = row[3]
        versioning = row[4]
        jalons_concernes = row[5]
        description = row[6]
        fmt = row[7] if len(row) > 7 else None

        if not isinstance(num, int) or not jalon or not document:
            continue

        jalon = str(jalon).strip()
        document = str(document).strip()

        base = slugify(f"{jalon}_{document}")
        count = seen_slugs.get(base, 0) + 1
        seen_slugs[base] = count
        code = base if count == 1 else f"{base}_{count}"

        entry = {
            "code": code,
            "num": num,
            "name": document,
            "jalon": jalon,
            "propriete": (propriete or "").strip() or "Informatif",
            "versioning": (versioning or "").strip() or "Document unique",
            "jalons_concernes": (str(jalons_concernes).strip() if jalons_concernes else jalon),
            # Champ `note` = description métier V12 (1 phrase). Nom conservé
            # pour compat avec matcher.py / types/juridique.py / l'UI.
            "note": (str(description).strip() if description else None) or None,
            "format": (str(fmt).strip() if fmt else None) or None,
        }
        docs_by_jalon.setdefault(jalon, []).append(entry)

    # Ordre canonique
    ordered_jalons: list[dict[str, Any]] = []
    for j in JALONS_ORDER:
        if j in docs_by_jalon:
            ordered_jalons.append({"jalon": j, "documents": docs_by_jalon[j]})
    # Jalons inattendus en fin
    for j, docs in docs_by_jalon.items():
        if j not in JALONS_ORDER:
            ordered_jalons.append({"jalon": j, "documents": docs})

    total = sum(len(j["documents"]) for j in ordered_jalons)

    result = {
        "version": "V12",
        "source": input_path.name,
        "total_documents": total,
        "jalons": ordered_jalons,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Convertit V12.xlsx → documents_v12.json")
    parser.add_argument("--in", dest="input", required=True, help="Chemin V12.xlsx")
    parser.add_argument("--out", dest="output", default="config/documents_v12.json")
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)
    if not in_path.exists():
        print(f"❌ Introuvable : {in_path}", file=sys.stderr)
        sys.exit(1)

    result = convert(in_path, out_path)
    print(f"✓ {result['total_documents']} documents convertis ({len(result['jalons'])} jalons)")
    print(f"✓ Écrit : {out_path}")
    for j in result["jalons"]:
        print(f"  - {j['jalon']:>15} : {len(j['documents'])} docs")


if __name__ == "__main__":
    main()
