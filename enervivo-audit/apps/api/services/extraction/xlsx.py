"""Extraction XLSX/XLSB via openpyxl — texte concaténé des cellules non vides.

Cas d'usage : TADD (Tableau Aide Décision), suivis budget, relevés
d'exploitation. On extrait jusqu'à `_HEAD_BUDGET` chars (sheet par sheet,
ligne par ligne) puis on coupe — suffisant pour identifier le doc.
"""

from __future__ import annotations

import io

from .base import HEAD_CHARS, ExtractionError, TextExtractor

_HEAD_BUDGET = HEAD_CHARS * 2


class XlsxExtractor(TextExtractor):
    def extract(self, content: bytes) -> str:
        try:
            from openpyxl import load_workbook  # déjà dans deps
        except ImportError as e:  # pragma: no cover
            raise ExtractionError(f"openpyxl absent : {e}") from e

        try:
            # read_only + data_only : rapide, pas d'eval de formules, suffisant
            # pour récupérer les valeurs visibles.
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            raise ExtractionError(f"openpyxl error : {e}") from e

        chunks: list[str] = []
        total = 0
        for sheet in wb.worksheets:
            chunks.append(f"=== {sheet.title} ===")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None and str(c).strip()]
                if cells:
                    line = " | ".join(cells)
                    chunks.append(line)
                    total += len(line)
                    if total >= _HEAD_BUDGET:
                        break
            if total >= _HEAD_BUDGET:
                break
        wb.close()

        text = "\n".join(chunks)
        if not text.strip():
            raise ExtractionError("XLSX vide")
        return text
